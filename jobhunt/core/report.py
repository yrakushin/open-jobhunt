from __future__ import annotations

import html
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from jobhunt.models import ApplicationResult, RunReport


def write_excel_report(report: RunReport, reports_dir: Path, candidate_label: str = "User") -> Path:
    reports_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    out = reports_dir / f"Otclicki_hh_{candidate_label}_{stamp}.xlsx"

    wb = Workbook()
    ws_sent = wb.active
    ws_sent.title = "Отклики"
    ws_sent.append(["статус", "должность", "компания", "регион", "зарплата", "match%", "ссылка", "текст письма", "время"])
    now = datetime.now().strftime("%H:%M")
    for item in report.sent:
        v = item.vacancy
        ws_sent.append([
            "отклик отправлен",
            v.title,
            v.company,
            v.region,
            v.salary,
            v.match_percent,
            v.url,
            item.letter,
            now,
        ])

    ws_manual = wb.create_sheet("Требует действия")
    ws_manual.append(["причина", "должность", "компания", "зарплата", "match%", "ссылка", "что сделать вручную", "заметки"])
    for item in report.manual:
        v = item.vacancy
        ws_manual.append([item.reason, v.title, v.company, v.salary, v.match_percent, v.url, item.notes, ""])

    ws_skip = wb.create_sheet("Пропущено")
    ws_skip.append(["причина", "должность", "компания", "зарплата", "match%", "ссылка"])
    for item in report.skipped:
        v = item.vacancy
        ws_skip.append([item.reason, v.title, v.company, v.salary, v.match_percent, v.url])

    wb.save(out)
    report.output_path = out
    return out
