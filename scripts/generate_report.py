#!/usr/bin/env python3
"""Generate weekly hh.ru autopilot Excel report."""
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

REPORTS = Path(__file__).resolve().parent.parent / "reports"


def main():
    data = json.loads(sys.stdin.read())
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M")
    out = REPORTS / f"Otclicki_hh_User_{ts}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Отклики"
    ws.append(["статус", "должность", "компания", "регион", "зарплата", "match%", "ссылка", "текст письма", "время"])
    for row in data.get("applied", []):
        ws.append([
            "отклик отправлен",
            row["title"],
            row["company"],
            row.get("region", "Москва"),
            row.get("salary", "не указана"),
            row.get("match", ""),
            row["url"],
            row.get("letter", ""),
            row.get("time", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ])

    ws2 = wb.create_sheet("Требует действия")
    ws2.append(["причина", "должность", "компания", "зарплата", "match%", "ссылка", "что сделать вручную", "заметки"])
    for row in data.get("action", []):
        ws2.append([
            row.get("reason", ""),
            row.get("title", ""),
            row.get("company", ""),
            row.get("salary", ""),
            row.get("match", ""),
            row.get("url", ""),
            row.get("todo", ""),
            row.get("notes", ""),
        ])

    ws3 = wb.create_sheet("Пропущено")
    ws3.append(["причина", "должность", "компания", "зарплата", "match%", "ссылка"])
    for row in data.get("skipped", []):
        ws3.append([
            row.get("reason", ""),
            row.get("title", ""),
            row.get("company", ""),
            row.get("salary", ""),
            row.get("match", ""),
            row.get("url", ""),
        ])

    wb.save(out)
    print(str(out))


if __name__ == "__main__":
    main()
